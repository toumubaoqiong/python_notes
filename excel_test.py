'''
Excel操作
'''

import openpyxl

def excel_operator():
    wb = openpyxl.load_workbook('example.xlsx')
    type(wb)
    # print(wb.get_sheet_names())
    sheet = wb.get_sheet_by_name('Sheet3')
    # print(sheet)
    # print(sheet.title)
    anotherSheet = wb.get_active_sheet()
    print(anotherSheet.title)


def cell_operator():
    wb = openpyxl.load_workbook('example.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    print(sheet["A1"])
    print(sheet["A1"].value)
    c = sheet["B1"]
    print(c.value)
    print("Row %s Column %s is %s" % (str(c.row), c.column, c.value))
    print("Cell %s is %s" % (c.coordinate, c.value))
    sheet.cell(row=1, column=2)
    print(sheet.cell(row=1, column=3).value)

    for i in range(1, 8, 2):
        print(sheet.cell(row=i, column=2).value)

    # print(sheet.get_highest_row())
    # print(sheet.get_highest_column())
    print(openpyxl.cell.get_column_letter(1))
    print(openpyxl.cell.column_index_from_string("A"))



# if __name__ == '__main__':
#     # excel_operator()
#     cell_operator()
